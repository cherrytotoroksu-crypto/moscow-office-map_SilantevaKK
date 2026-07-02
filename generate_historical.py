"""
Universal script to generate buildings_QQQQMM.json and lots_QQQQMM.json
for all historical quarters.
"""
import sys, json, os, warnings
warnings.filterwarnings('ignore')
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

DATA_DIR = r'C:\Users\zapas\OneDrive\Documents\курсач\moscow-office-map\data'
TD = r'C:\Users\zapas\Downloads\Telegram Desktop'

# Бадаевский split params (same as other quarters)
BADAEVSKY_W = {'name': 'Бадаевский Западная лента', 'lat': 55.749324, 'lng': 37.558800, 'gba': 7713, 'year': 2026}
BADAEVSKY_E = {'name': 'Бадаевский Восточная лента', 'lat': 55.749324, 'lng': 37.560200, 'gba': 11145, 'year': 2027}

def save_json(data, path):
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f'  Saved: {os.path.basename(path)}')

def clean(v):
    if v is None: return ''
    return str(v).strip()

def to_float(v, default=0.0):
    try: return float(v) if v else default
    except: return default

def to_int(v, default=0):
    try: return int(float(v)) if v else default
    except: return default

def normalize_cls(v):
    s = clean(v).upper()
    if s in ('A', 'А'): return 'A'
    if s in ('B', 'В', 'B+'): return 'B+'
    if s in ('B-', 'В-'): return 'B-'
    return s or 'B+'

def normalize_zone(v):
    s = clean(v)
    # Fix Север.Сити zone
    if 'БК-СК' in s or s == 'БК': return 'ТТК-МКАД'
    return s

def badaevsky_split(b_template, volume_total, price):
    """Split Бадаевский into West and East entries."""
    result = []
    for spec in [BADAEVSKY_W, BADAEVSKY_E]:
        b = dict(b_template)
        b['name'] = spec['name']
        b['name_orig'] = b_template.get('name_orig', 'Бадаевский')
        b['lat'] = spec['lat']
        b['lng'] = spec['lng']
        b['gba'] = spec['gba']
        b['year'] = spec['year']
        # Split volume proportionally by gba
        total_gba = BADAEVSKY_W['gba'] + BADAEVSKY_E['gba']
        b['volume'] = round(volume_total * spec['gba'] / total_gba, 1)
        b['price'] = price
        b['lots'] = 0
        b['on_sale'] = 'нет'
        result.append(b)
    return result

def read_lots_2024(ws_blocks, bldg_name_col=0, block_col=1, floor_col=2,
                   area_col=3, price_col=4, scheme_col=6, size_col=None):
    """Parse lots from 2024-era блоки sheet."""
    lots = {}
    for row in ws_blocks.iter_rows(values_only=True):
        if not row[bldg_name_col] or str(row[bldg_name_col]).strip() in ('Здание', ''):
            continue
        bname = str(row[bldg_name_col]).strip()
        area = to_float(row[area_col])
        price = to_float(row[price_col])
        if area <= 0: continue
        # determine size
        if size_col is not None:
            size = clean(row[size_col])
        else:
            # try last non-None
            size = ''
            for i in range(len(row)-1, -1, -1):
                if row[i] and str(row[i]).strip() not in ('', 'None'):
                    size = str(row[i]).strip()
                    break
        lot = {
            'block': clean(row[block_col]),
            'floor': to_int(row[floor_col]),
            'area': round(area, 1),
            'price': round(price),
            'total': round(area * price),
            'scheme': clean(row[scheme_col]),
            'size': size,
        }
        lots.setdefault(bname, []).append(lot)
    return lots

def read_lots_2023(ws_blocks):
    """Parse lots from 2023-era Блоки sheet.
    Header: Здание(0), Название блока(1), Этаж(2), Площадь(3), Цена(4), Вес(5),
            Схема(6), Здание/Блок(7), Субрынок(8), Размер(9), Класс(10)
    """
    lots = {}
    for row in ws_blocks.iter_rows(values_only=True):
        if not row[0] or str(row[0]).strip() in ('Здание', ''):
            continue
        bname = str(row[0]).strip()
        area = to_float(row[3])
        price = to_float(row[4])
        if area <= 0: continue
        size = clean(row[9]) if len(row) > 9 else ''
        lot = {
            'block': clean(row[1]),
            'floor': to_int(row[2]),
            'area': round(area, 1),
            'price': round(price),
            'total': round(area * price),
            'scheme': clean(row[6]),
            'size': size,
        }
        lots.setdefault(bname, []).append(lot)
    return lots

def read_lots_2021(ws_blocks):
    """Parse lots from 2021-era Блоки sheet.
    May 2021: ...Зона(10), Субрынок(11), Класс(12), Размер(13)
    Dec 2021: ...Здание/Блок(9), Размер(10), Класс(11), Субрынок(12)
    Dec 2022: ...Схема(8), Класс(9), Здание/Блок(10), Размер(11), Субрынок(12)
    """
    # Detect layout from header
    header = None
    rows_iter = ws_blocks.iter_rows(values_only=True)
    size_col = 13  # default for May 2021
    price_vat_col = 5
    scheme_col = 8
    for row in rows_iter:
        if row[0] and str(row[0]).strip() == 'Здание':
            header = row
            # Detect size column by scanning header
            for i, h in enumerate(header):
                if h and 'размер' in str(h).lower():
                    size_col = i
                    break
            # Detect if has explicit price_vat col
            for i, h in enumerate(header):
                if h and 'цена с ндс' in str(h).lower().replace(' ', ' '):
                    price_vat_col = i
                    break
            break

    lots = {}
    for row in rows_iter:
        if not row[0] or str(row[0]).strip() in ('Здание', ''):
            continue
        bname = str(row[0]).strip()
        area = to_float(row[3])
        price_vat = to_float(row[price_vat_col])
        if area <= 0 or price_vat <= 0: continue
        size = clean(row[size_col]) if len(row) > size_col else ''
        lot = {
            'block': clean(row[1]),
            'floor': to_int(row[2]),
            'area': round(area, 1),
            'price': round(price_vat),
            'total': round(area * price_vat),
            'scheme': clean(row[scheme_col]),
            'size': size,
        }
        lots.setdefault(bname, []).append(lot)
    return lots

def parse_202212(excel_path, qid):
    """December 2022: Build sheet with 20 cols.
    name(0), addr(1), dev(2), start(3), status(4), year(5), class(6), gba(7),
    volume(8), price_vat(9), wt_vat(10), price_novat(11), wt(12), sold(13),
    reserve(14), district(15), zone(16), submarket(17), lat(18), lng(19)
    Блоки: name(0), block(1), floor(2), area(3), price_novat(4), price_vat(5),
           weight(6), wt_vat(7), scheme(8), class(9), type(10), size(11)
    """
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb['Build']
    buildings = []

    for row in ws.iter_rows(values_only=True):
        if not row[0] or str(row[0]).strip() in ('Название здания', ''):
            continue
        name = clean(row[0])
        status = clean(row[4]).lower()
        lat = to_float(row[18])
        lng = to_float(row[19])
        if lat == 0 or lng == 0:
            print(f'  SKIP (no coords): {name}')
            continue
        on_sale = 'да' if status in ('строится', 'строится/построен', 'построен') else 'нет'
        price = to_float(row[9])
        volume = to_float(row[8])
        zone = normalize_zone(clean(row[16]))

        b = {
            'name': name,
            'name_orig': name,
            'lat': round(lat, 6),
            'lng': round(lng, 6),
            'volume': round(volume, 1),
            'price': round(price),
            'cls': normalize_cls(row[6]),
            'status': clean(row[4]),
            'on_sale': on_sale,
            'lots': 0,
            'gba': to_int(row[7]),
            'gla': to_int(row[7]),
            'zone': zone,
            'submarket': clean(row[17]),
            'biz': '',
            'developer': clean(row[2]),
            'address': clean(row[1]),
            'year': clean(row[5]),
            'start': clean(row[3]),
            'weight': to_float(row[12]),
            'id': name.lower().replace(' ', '_'),
        }
        if 'бадаевский' in name.lower():
            parts = badaevsky_split(b, volume, price)
            buildings.extend(parts)
        else:
            buildings.append(b)

    try:
        ws_b = wb['Блоки']
        lots = read_lots_2021(ws_b)
        for b in buildings:
            b['lots'] = len(lots.get(b['name'], []))
    except Exception as e:
        print(f'  No Блоки: {e}')
        lots = {}

    wb.close()
    return buildings, lots

# ─── FORMAT 2024+ ────────────────────────────────────────────────────────────
def parse_2024plus(excel_path, qid):
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb['здания']
    buildings = []
    for row in ws.iter_rows(values_only=True):
        if not row[0] or str(row[0]).strip() in ('Название здания', ''):
            continue
        name = clean(row[0])
        status = clean(row[4]).lower()
        lat = to_float(row[16])
        lng = to_float(row[17])
        if lat == 0 or lng == 0:
            print(f'  SKIP (no coords): {name}')
            continue
        on_sale = 'да' if status in ('строится', 'строится/построен', 'построен') else 'нет'
        price = to_float(row[10])
        volume = to_float(row[9])
        zone = normalize_zone(clean(row[14]))

        b = {
            'name': name,
            'name_orig': name,
            'lat': round(lat, 6),
            'lng': round(lng, 6),
            'volume': round(volume, 1),
            'price': round(price),
            'cls': normalize_cls(row[6]),
            'status': clean(row[4]),
            'on_sale': on_sale,
            'lots': 0,
            'gba': to_int(row[7]),
            'gla': to_int(row[7]),
            'zone': zone,
            'submarket': clean(row[15]),
            'biz': '',
            'developer': clean(row[2]),
            'address': clean(row[1]),
            'year': clean(row[5]),
            'start': clean(row[3]),
            'weight': to_float(row[11]),
            'id': name.lower().replace(' ', '_'),
        }
        # Бадаевский split
        if 'бадаевский' in name.lower():
            parts = badaevsky_split(b, volume, price)
            buildings.extend(parts)
        else:
            buildings.append(b)

    # Read lots
    try:
        ws_b = wb['блоки']
        # Detect column layout: check if col 8 is Субрынок or Здание/Блок
        hdr = next(ws_b.iter_rows(values_only=True))
        if hdr[8] and 'субрынок' in str(hdr[8]).lower():
            # 202403 layout: size at col 10
            lots = read_lots_2024(ws_b, size_col=10)
        else:
            # 202412/202410 layout: size at col 9
            lots = read_lots_2024(ws_b, size_col=9)
        # attach lot counts
        for b in buildings:
            b['lots'] = len(lots.get(b['name'], []) + lots.get(b.get('name_orig', ''), []))
    except Exception as e:
        print(f'  No блоки sheet: {e}')
        lots = {}

    wb.close()
    return buildings, lots

# ─── FORMAT 2023 (Build sheet) ───────────────────────────────────────────────
def parse_2023(excel_path, qid, has_zhiloy=False):
    """has_zhiloy=True for 202308 which has an extra 'Жилой?' column."""
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb['Build']
    buildings = []
    shift = 1 if has_zhiloy else 0  # column shift due to 'Жилой?' extra col

    for row in ws.iter_rows(values_only=True):
        if not row[0] or str(row[0]).strip() in ('Название здания', ''):
            continue
        name = clean(row[0])
        developer = clean(row[2 + shift])
        status = clean(row[4 + shift]).lower()
        year = clean(row[5 + shift])
        cls = normalize_cls(row[6 + shift])
        gba = to_int(row[7 + shift])
        volume = to_float(row[8 + shift])
        price = to_float(row[9 + shift])
        zone = normalize_zone(clean(row[13 + shift]))
        submarket = clean(row[14 + shift])
        lat = to_float(row[15 + shift])
        lng = to_float(row[16 + shift])

        if lat == 0 or lng == 0:
            print(f'  SKIP (no coords): {name}')
            continue
        on_sale = 'да' if status in ('строится', 'строится/построен', 'построен') else 'нет'

        b = {
            'name': name,
            'name_orig': name,
            'lat': round(lat, 6),
            'lng': round(lng, 6),
            'volume': round(volume, 1),
            'price': round(price),
            'cls': cls,
            'status': clean(row[4 + shift]),
            'on_sale': on_sale,
            'lots': 0,
            'gba': gba,
            'gla': gba,
            'zone': zone,
            'submarket': submarket,
            'biz': '',
            'developer': developer,
            'address': clean(row[1]),
            'year': year,
            'start': clean(row[3 + shift]),
            'weight': to_float(row[10 + shift]),
            'id': name.lower().replace(' ', '_'),
        }
        if 'бадаевский' in name.lower():
            parts = badaevsky_split(b, volume, price)
            buildings.extend(parts)
        else:
            buildings.append(b)

    # Read lots
    try:
        ws_b = wb['Блоки']
        lots = read_lots_2023(ws_b)
        for b in buildings:
            b['lots'] = len(lots.get(b['name'], []) + lots.get(b.get('name_orig', ''), []))
    except Exception as e:
        print(f'  No Блоки sheet: {e}')
        lots = {}

    wb.close()
    return buildings, lots

# ─── FORMAT 2021-2022 (Здания sheet) ─────────────────────────────────────────
def parse_2021_2022(excel_path, qid):
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb['Здания']
    buildings = []

    for row in ws.iter_rows(values_only=True):
        if not row[0] or str(row[0]).strip() in ('Название здания', ''):
            continue
        name = clean(row[0])
        # col 4 = year (no status column)
        year = clean(row[4])
        cls = normalize_cls(row[5])
        gba = to_int(row[6])
        volume = to_float(row[7])
        price_no_vat = to_float(row[8])
        price_vat = round(price_no_vat * 1.2)  # convert to с НДС
        zone = normalize_zone(clean(row[12]))
        submarket = clean(row[13])
        lat = to_float(row[14])
        lng = to_float(row[15])

        if lat == 0 or lng == 0:
            print(f'  SKIP (no coords): {name}')
            continue

        b = {
            'name': name,
            'name_orig': name,
            'lat': round(lat, 6),
            'lng': round(lng, 6),
            'volume': round(volume, 1),
            'price': price_vat,
            'cls': cls,
            'status': 'Строится',
            'on_sale': 'да',
            'lots': 0,
            'gba': gba,
            'gla': gba,
            'zone': zone,
            'submarket': submarket,
            'biz': '',
            'developer': clean(row[2]),
            'address': clean(row[1]),
            'year': year,
            'start': clean(row[3]),
            'weight': to_float(row[9]),
            'id': name.lower().replace(' ', '_'),
        }
        if 'бадаевский' in name.lower():
            parts = badaevsky_split(b, volume, price_vat)
            buildings.extend(parts)
        else:
            buildings.append(b)

    # Read lots
    try:
        ws_b = wb['Блоки']
        lots = read_lots_2021(ws_b)
        for b in buildings:
            b['lots'] = len(lots.get(b['name'], []) + lots.get(b.get('name_orig', ''), []))
    except Exception as e:
        print(f'  No Блоки sheet: {e}')
        lots = {}

    wb.close()
    return buildings, lots

# ─── FORMAT 202407 (карта sheet only) ────────────────────────────────────────
def parse_202407(excel_path, qid):
    """карта sheet: name(0), addr(1), year(2), class(3), gba(4), volume(5), price(6), submarket(7)
    No coordinates, no lots, no developer."""
    # Build coord/developer lookup from existing JSON files
    coord_lookup = {}
    for fname in os.listdir(DATA_DIR):
        if fname.startswith('buildings_') and fname.endswith('.json'):
            with open(os.path.join(DATA_DIR, fname), encoding='utf-8') as f:
                try:
                    data = json.load(f)
                    for b in data:
                        n = b.get('name', '').strip().lower()
                        if n and b.get('lat') and b.get('lng'):
                            coord_lookup[n] = {
                                'lat': b['lat'], 'lng': b['lng'],
                                'developer': b.get('developer', ''),
                                'zone': b.get('zone', ''),
                            }
                except: pass

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb['карта']
    buildings = []

    for row in ws.iter_rows(values_only=True):
        if not row[0] or str(row[0]).strip() in ('Название здания', ''):
            continue
        name = clean(row[0])
        key = name.lower()
        coords = coord_lookup.get(key, {})
        lat = coords.get('lat', 0)
        lng = coords.get('lng', 0)
        if lat == 0 or lng == 0:
            print(f'  SKIP (no coords in lookup): {name}')
            continue

        b = {
            'name': name,
            'name_orig': name,
            'lat': lat,
            'lng': lng,
            'volume': round(to_float(row[5]), 1),
            'price': round(to_float(row[6])),
            'cls': normalize_cls(row[3]),
            'status': 'Строится',
            'on_sale': 'да',
            'lots': 0,
            'gba': to_int(row[4]),
            'gla': to_int(row[4]),
            'zone': normalize_zone(coords.get('zone', '')),
            'submarket': clean(row[7]),
            'biz': '',
            'developer': coords.get('developer', ''),
            'address': clean(row[1]),
            'year': clean(row[2]),
            'start': '',
            'weight': 0,
            'id': name.lower().replace(' ', '_'),
        }
        buildings.append(b)

    wb.close()
    return buildings, {}

# ─── QUARTERS TO GENERATE ────────────────────────────────────────────────────
QUARTERS = [
    ('202105', os.path.join(TD, 'Свод_рабочая_Май 2021.xlsx'), '2021'),
    ('202112', os.path.join(TD, 'Свод_рабочая_Декабрь 2021.xlsx'), '2021'),
    ('202207', os.path.join(TD, 'Свод_рабочая_Июль 2022.xlsx'), '2021'),
    ('202212', os.path.join(TD, 'Свод_рабочая_Декабрь 2022.xlsx'), '202212'),
    ('202305', os.path.join(TD, 'Свод_рабочая_Май 2023.xlsx'), '2023'),
    ('202308', os.path.join(TD, 'Свод_рабочая_Август 2023.xlsx'), '2023_zhiloy'),
    ('202311', os.path.join(TD, 'Свод_рабочая_Ноябрь_2023_AB_обновленный_STONE.xlsx'), '2023'),
    ('202403', os.path.join(TD, 'Свод_рабочая_Март 2024.xlsx'), '2024'),
    ('202407', os.path.join(TD, 'Свод_рабочая_Июль 2024_AB.xlsx'), '202407'),
]

for qid, path, era in QUARTERS:
    print(f'\n=== {qid} ({era}) ===')
    try:
        if era == '2021':
            buildings, lots = parse_2021_2022(path, qid)
        elif era == '202212':
            buildings, lots = parse_202212(path, qid)
        elif era == '2023':
            buildings, lots = parse_2023(path, qid, has_zhiloy=False)
        elif era == '2023_zhiloy':
            buildings, lots = parse_2023(path, qid, has_zhiloy=True)
        elif era == '2024':
            buildings, lots = parse_2024plus(path, qid)
        elif era == '202407':
            buildings, lots = parse_202407(path, qid)
        else:
            print(f'  Unknown era: {era}')
            continue

        print(f'  Buildings: {len(buildings)}')
        print(f'  Lots dict: {len(lots)} buildings with lots')

        bldg_path = os.path.join(DATA_DIR, f'buildings_{qid}.json')
        lots_path = os.path.join(DATA_DIR, f'lots_{qid}.json')
        save_json(buildings, bldg_path)
        save_json(lots, lots_path)
    except Exception as e:
        import traceback
        print(f'  ERROR: {e}')
        traceback.print_exc()

print('\nDone.')
